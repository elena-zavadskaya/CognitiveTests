from turtle import pd

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from .serializers import ProfileSerializer, ResultSerializer, TestSerializer
from django.contrib.auth.hashers import make_password, check_password
import logging

import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from .models import AppUser, Profile, Result, Test

logger = logging.getLogger(__name__)

class RegisterView(APIView):
    def post(self, request):
        logger.info("Полученные данные при регистрации: %s", request.data)
        nickname = request.data.get('nickname')
        password = request.data.get('password')

        if not nickname or not password:
            return Response({'error': 'Необходимо указать никнейм и пароль'}, status=status.HTTP_400_BAD_REQUEST)

        hashed_password = make_password(password)
        logger.info("Хэшированный пароль: %s", hashed_password)

        user = AppUser.objects.create(nickname=nickname, password_hash=hashed_password)
        return Response({'message': 'Пользователь успешно зарегистрирован', 'user_id': user.id}, status=status.HTTP_201_CREATED)

class LoginView(APIView):
    def post(self, request):
        nickname = request.data.get('nickname')
        password = request.data.get('password')

        if not nickname or not password:
            return Response({'error': 'Необходимо указать никнейм и пароль'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = AppUser.objects.get(nickname=nickname)
        except AppUser.DoesNotExist:
            return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

        if check_password(password, user.password_hash):
            return Response({'message': 'Успешный вход', 'user_id': user.id}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Неверный пароль'}, status=status.HTTP_400_BAD_REQUEST)

class ProfileView(APIView):
    def post(self, request):
        logger.info("Полученные данные: %s", request.data)
        user_id = request.data.get('user')
        if not user_id:
            return Response({"error": "User ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profile = Profile.objects.get(user_id=user_id)
            serializer = ProfileSerializer(profile, data=request.data, partial=True)
        except Profile.DoesNotExist:
            serializer = ProfileSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        logger.error("Ошибки валидации: %s", serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ResultView(APIView):
    def post(self, request):
        logger.info("Полученные данные: %s", request.data)

        # Преобразование данных
        request.data['score_percentage'] = float(request.data.get('score_percentage', 0))
        request.data['number_all_answers'] = int(request.data.get('number_all_answers', 0))
        request.data['number_correct_answers'] = int(request.data.get('number_correct_answers', 0))

        serializer = ResultSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        logger.error("Ошибки валидации: %s", serializer.errors)
        return Response(
            {"error": "Invalid data", "details": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

class TestListView(APIView):
    def get(self, request):
        tests = Test.objects.all()
        serializer = TestSerializer(tests, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TestDetailView(APIView):
    def get(self, request, test_id):
        try:
            test = Test.objects.get(id=test_id)
            serializer = TestSerializer(test)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Test.DoesNotExist:
            return Response({'error': 'Тест не найден'}, status=status.HTTP_404_NOT_FOUND)


def export_users_results_excel(request):
    users = AppUser.objects.all().select_related('profile')
    all_tests = Test.objects.order_by('id')

    base_columns = [
        "ИД", "Возраст", "Образование", "Специальность",
        "Место жительства", "Рост", "Вес", "Ведущая рука", "Заболевания",
        "Курение", "Алкоголь", "Спорт", "Бессонница", "Текущее состояние", "Геймер"
    ]

    test_headers = {test.title: ["% выполнения", "Время выполнения", "Количество правильных ответов", "Количество вопросов"]
                    for test in all_tests}

    all_columns = base_columns + [f"{header} ({test})" for test in test_headers for header in test_headers[test]]

    data = []
    for user in users:
        profile = getattr(user, 'profile', None)
        user_results = Result.objects.filter(user=user).select_related('test')

        test_attempts = {test.title: [] for test in all_tests}
        for result in user_results:
            test_attempts[result.test.title].append(result)

        max_attempts = max([len(attempts) for attempts in test_attempts.values()], default=1)

        for attempt_idx in range(max_attempts):
            row = {col: "" for col in all_columns}
            if attempt_idx == 0:
                row.update({
                    "ИД": user.id,
                    "Возраст": profile.age if profile else "",
                    "Образование": profile.education if profile else "",
                    "Специальность": profile.speciality if profile else "",
                    "Место жительства": profile.residence if profile else "",
                    "Рост": profile.height if profile else "",
                    "Вес": profile.weight if profile else "",
                    "Ведущая рука": profile.lead_hand if profile else "",
                    "Заболевания": profile.diseases if profile else "",
                    "Курение": "Да" if profile and profile.smoking else "Нет",
                    "Алкоголь": profile.alcohol if profile else "",
                    "Спорт": profile.sport if profile else "",
                    "Бессонница": "Да" if profile and profile.insomnia else "Нет",
                    "Текущее состояние": profile.current_health if profile else "",
                    "Геймер": "Да" if profile and profile.gaming else "Нет",
                })

            for test in all_tests:
                attempts = test_attempts.get(test.title, [])
                if attempt_idx < len(attempts):
                    row[f"% выполнения ({test.title})"] = attempts[attempt_idx].score_percentage
                    row[f"Время выполнения ({test.title})"] = attempts[attempt_idx].time
                    row[f"Количество правильных ответов ({test.title})"] = attempts[attempt_idx].number_correct_answers
                    row[f"Количество вопросов ({test.title})"] = attempts[attempt_idx].number_all_answers

            data.append(row)

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users_results.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    font_style = Font(name="Times New Roman", size=12)
    header_font = Font(name="Times New Roman", size=12, bold=True)

    col_index = len(base_columns) + 1
    for test, subheaders in test_headers.items():
        ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + len(subheaders) - 1)
        cell = ws.cell(row=1, column=col_index, value=test)
        cell.border = thin_border
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_index += len(subheaders)

    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title.split(" (")[0])
        cell.border = thin_border
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_index = 3
    prev_user_row = row_index

    for i, row in df.iterrows():
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col_num, value=value)
            cell.border = thin_border
            cell.font = font_style
            cell.alignment = Alignment(vertical="center")

        if row["ИД"] != "":
            prev_user_row = row_index

        if (row_index + 1 > len(df)) or (df.iloc[i + 1]["ИД"] != ""):
            for col_num in range(1, len(base_columns) + 1):
                ws.merge_cells(start_row=prev_user_row, start_column=col_num, end_row=row_index, end_column=col_num)
                ws.cell(row=prev_user_row, column=col_num).alignment = Alignment(vertical="center", horizontal="center")

        row_index += 1

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        if cell.value:
            ws.column_dimensions[cell.column_letter].width = 25

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 35

    wb.save(response)
    return response

